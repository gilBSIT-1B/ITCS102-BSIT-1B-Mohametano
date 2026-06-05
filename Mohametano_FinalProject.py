from tkinter import *
import openpyxl as op
from tkinter import messagebox, ttk
import os

window = Tk()#to call a window
window.title("[Scholarship Applicant Tracking System]")#Title malamang
window.config(background="skyblue")#pagbabago ng background color

title = Label(window, text="Scholarship Applicant Tracking System", 
                 font=("arial", 20, "bold"), bg="skyblue")
title.grid(row=0, column=0, columnspan=8, pady=(20))
#==================================================================def functions
def create_excel_db():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists("Mohametano_Database.xlsx"):
        wb = op.Workbook()
        ws = wb.active
        ws.title = "Scholarship Applicants"
        ws.append(["ID", "Applicant Name", "Sex", "Age", 
           "Civil Status", "Home Address", "School name", "Email Address", "Phone Number"])
        
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        
        wb.save("Mohametano_Database.xlsx")
create_excel_db()

def display_excel():
    workbook = op.load_workbook('Mohametano_Database.xlsx')
    sheet = workbook.active

    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            table.insert("",END, values=row)
            
def validation_input():
    name = name_entry.get()
    sex = Sex_lbox.get()
    age = age_entry.get()
    status = status_lbox.get()
    home = home_entry.get()
    school = school_entry.get()
    email = email_entry.get()
    phone = phone_entry.get()

    if not name or not sex or not age or not status or not home or not school or not email or not phone:
        messagebox.showerror("ERROR", "All fields are required!")
        return False

    if not age.isdigit():
        messagebox.showerror("ERROR", "Age must be a number!")
        return False

    if not phone.isdigit():
        messagebox.showerror("ERROR", "Phone number must contain only numbers!")
        return False

    if len(phone) == 11:
        messagebox.showerror("ERROR", "Phone number must be 11 digits!")
        return False

    return True

def append_excel():
    if not validation_input():
        return

    workbook = op.load_workbook("Mohametano_Database.xlsx")
    sheet = workbook.active

    new_id = sheet.max_row

    sheet.append([
        new_id,
        name_entry.get(),
        Sex_lbox.get(),
        age_entry.get(),
        status_lbox.get(),
        home_entry.get(),
        school_entry.get(),
        email_entry.get(),
        phone_entry.get()
    ])

    workbook.save("Mohametano_Database.xlsx")

    messagebox.showinfo("SUCCESS", "Record Added Successfully!")

    display_excel()

def update_record():
    selected = table.focus()

    if not selected:
        messagebox.showerror("ERROR", "Select a record first!")
        return

    if not validation_input():
        return

    record_id = table.item(selected)["values"][0]

    workbook = op.load_workbook("Mohametano_Database.xlsx")
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == record_id:

            sheet.cell(row, 2).value = name_entry.get()
            sheet.cell(row, 3).value = Sex_lbox.get()
            sheet.cell(row, 4).value = age_entry.get()
            sheet.cell(row, 5).value = status_lbox.get()
            sheet.cell(row, 6).value = home_entry.get()
            sheet.cell(row, 7).value = school_entry.get()
            sheet.cell(row, 8).value = email_entry.get()
            sheet.cell(row, 9).value = phone_entry.get()

            break

    workbook.save("Mohametano_Database.xlsx")

    display_excel()
    clear_entries()

    messagebox.showinfo("SUCCESS", "Record Updated!")

def clear_entries():
    name_entry.delete(0, END)
    Sex_lbox.set("")
    age_entry.delete(0, END)
    status_lbox.set("")
    home_entry.delete(0, END)
    school_entry.delete(0, END)
    email_entry.delete(0, END)
    phone_entry.delete(0, END)

    table.selection_remove(table.selection())

def select_record(event):
    selected = table.focus()

    if selected:
        values = table.item(selected, "values")

        clear_entries()

        name_entry.insert(0, values[1])
        Sex_lbox.set(values[2])
        age_entry.insert(0, values[3])
        status_lbox.set(values[4])
        home_entry.insert(0, values[5])
        school_entry.insert(0, values[6])
        email_entry.insert(0, values[7])
        phone_entry.insert(0, values[8])


def delete_record():
    selected = table.focus()

    if not selected:
        messagebox.showerror("ERROR", "Select a record first!")
        return

    confirm = messagebox.askyesno(
        "DELETE",
        "Are you sure you want to delete this record?"
    )

    if not confirm:
        return

    record_id = table.item(selected)["values"][0]

    workbook = op.load_workbook("Mohametano_Database.xlsx")
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == record_id:
            sheet.delete_rows(row)
            break

    workbook.save("Mohametano_Database.xlsx")

    display_excel()
    clear_entries()

    messagebox.showinfo("SUCCESS", "Record Deleted!")

columns = ("ID", "Applicant Name", "Sex", "Age", 
           "Civil Status", "Home Address", "School name", "Email Address", "Phone Number")

table = ttk.Treeview(window, columns=columns, show="headings")
for col in columns:
    table.heading(col, text=col)
    if col == "ID":
        table.column(col, width=50, anchor=CENTER)
    elif col == "Applicant Name":
        table.column(col, width=130, anchor=W)
    elif col == "Sex":
        table.column(col, width=100, anchor=CENTER)
    elif col == "Age":
        table.column(col, width=90, anchor=CENTER)
    elif col == "Civil Status":
        table.column(col, width=100, anchor=CENTER)
    elif col == "Home Address":
        table.column(col, width=110, anchor=W)
    elif col == "Email Address":
        table.column(col, width=70, anchor=CENTER)
    elif col == "Phone Number":
        table.column(col, width=90, anchor=CENTER)
    else:
        table.column(col, width=130, anchor=CENTER)

table.grid(row=6, column=0, columnspan=8, padx=10, pady=(0, 10), sticky="nsew")

#======================================================================================================
name_label = Label(window, text="Name",font=("Arial", 15, "bold"))                      #Name LABEL
name_label.grid(row=1,column=0)                                 

name_entry = Entry(window,font=("Arial", 15, "bold"),)
name_entry.grid(row=1,column=1,padx=(0,10))                                             #Name entry
#=======================================================================================================

#=======================================================================================================
Sex_label = Label(window, text="Sex", font=("Arial", 15, "bold"))
Sex_label.grid(row=1,column=2)#                                                       Sex

Sex_lbox = ttk.Combobox(window, values=["Male", "Female"], state="readonly",font=("Arial", 15, "bold"))
Sex_lbox.grid(row=1,column=3,padx=(0,10))    
#=======================================================================================================

#=======================================================================================================
agelabel = Label(window,text="Age", font=("Arial", 15, "bold"))
agelabel.grid(row=1,column=4)   
#                                                                                                   AGE
age_entry = Entry(window,font=("Arial", 15, "bold"))
age_entry.grid(row=1, column=5,padx=(0,10))
#=======================================================================================================

#=======================================================================================================
status_label = Label(window,text="Civil Status", font=("Arial", 15, "bold"))
status_label.grid(row=1,column=6)
#                                                                                          MIRRAGE STATUS
status_lbox = ttk.Combobox(window, values=["Single", "Married"], state="readonly",font=("Arial", 15, "bold"))
status_lbox.grid(row=1, column=7,padx=(0,10))
#=======================================================================================================

#=====================================================================================================
home_label = Label(window,text="Home Address", font=("Arial", 15, "bold"))
home_label.grid(row=2, column=0, pady=25)
#                                                                                          HOME ADRESS
home_entry = Entry(window,font=("Arial", 15, "bold"))
home_entry.grid(row=2, column=1,padx=(0,10))
#=======================================================================================================

#=======================================================================================================
school_label = Label(window,text="School Name", font=("Arial", 15, "bold"))
school_label.grid(row=2, column=2, pady=25)
#                                                                                          SCHOOL NAME
school_entry = Entry(window,font=("Arial", 15, "bold"))
school_entry.grid(row=2, column=3,padx=(0,10))
#=======================================================================================================

#=======================================================================================================
email_label = Label(window,text="Email Address", font=("Arial", 15, "bold"))
email_label.grid(row=2, column=4, pady=25)
#                                                                                          EMAIL ADRESS
email_entry = Entry(window,font=("Arial", 15, "bold"))
email_entry.grid(row=2, column=5,padx=(0,10))
#=======================================================================================================

#=======================================================================================================
phone_label = Label(window,text="Phone Number", font=("Arial", 15, "bold"))
phone_label.grid(row=2, column=6, pady=25)
#                                                                                          PHONE NUMBER
phone_entry = Entry(window,font=("Arial", 15, "bold"))
phone_entry.grid(row=2, column=7,padx=(0,10))
#=======================================================================================================
label = Label(window,text='REGISTERED APPLICANTS', font=("Arial", 15, "bold"), bg='lightblue')
label.grid(row=4, column=0, columnspan=8)
#=======================================================================================================
#BUTTONS
submit_butt = Button(window, text="SUBMIT THIS RECORD", font=("Arial", 15, "bold"), bg="blue", fg="white", command=append_excel)
submit_butt.grid(row=3, column=0, columnspan=2)

clears_butt = Button(window, text="CLEAR ENTRIES", font=("Arial", 15, "bold"), bg="blue", fg="white", command=clear_entries)
clears_butt.grid(row=3, column=2, columnspan=2)

update_butt = Button(window, text="UPDATE A RECORD", font=("Arial", 15, "bold"), bg="blue", fg="white", command=update_record)
update_butt.grid(row=3, column=4, columnspan=2)

delete_butt = Button(window, text="REMOVE RECORD", font=("Arial", 15, "bold"), bg="blue", fg="white", command=delete_record)
delete_butt.grid(row=3, column=6, columnspan=2)

table.bind("<<TreeviewSelect>>", select_record)

display_excel()
window.mainloop()#To show the window